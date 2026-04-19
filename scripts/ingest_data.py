import argparse
import pandas as pd
from sqlalchemy import create_engine
import os
import re
import logging
import sys
from secrets_manager import get_database_url
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb

# Configure standard logging
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger(__name__)

DB_URL = get_database_url()


def get_engine():
    """Create database engine."""
    return create_engine(DB_URL)


def read_excel_with_dynamic_header(
    file_path,
    engine="openpyxl",
    sheet_name=None,
    keywords=[
        "Código de la Institución",
        "IES PADRE",
        "Institución de Educación Superior (IES)",
        "Sector IES",
        "ID Caracter",
    ],
):
    """
    Dynamically identifies the header row by searching for keywords.
    Skips metadata and empty rows.
    """
    # Read first 20 rows without header to find the actual header row
    if sheet_name:
        df_preview = pd.read_excel(
            file_path, engine=engine, header=None, nrows=20, sheet_name=sheet_name
        )
    else:
        df_preview = pd.read_excel(file_path, engine=engine, header=None, nrows=20)

    header_idx = None
    upper_keywords = [kw.upper() for kw in keywords]
    for i, row in df_preview.iterrows():
        # Clean row values for matching (handle partial matches and case insensitivity)
        row_values = [str(val).strip().upper() for val in row if pd.notna(val)]
        if any(any(kw in val for kw in upper_keywords) for val in row_values):
            header_idx = i
            break

    if header_idx is not None:
        logger.info(
            f"Dynamic header detected at row {header_idx} for {os.path.basename(file_path)}"
        )
        # Re-read the file starting from the detected header row (which becomes index 0)
        df = None
        if sheet_name:
            df = pd.read_excel(
                file_path,
                engine=engine,
                sheet_name=sheet_name,
                skiprows=header_idx,
            )
        else:
            df = pd.read_excel(file_path, engine=engine, skiprows=header_idx)
        return df

    logger.warning(
        f"Could not dynamically identify header for {os.path.basename(file_path)}. Falling back to default."
    )
    return pd.read_excel(file_path, engine=engine)


def format_columns(df):
    """Normalize dataframe columns to standard snake_case."""
    df.columns = [
        re.sub(r"\W+", "_", str(col).strip().lower()).strip("_") for col in df.columns
    ]
    return df


def load_file(file_path):
    """Loads a single file into the bronze schema."""
    engine = get_engine()
    file_name = os.path.basename(file_path)

    logger.info(f"Opening file and loading {file_name} into Bronze schema...")

    # Determine table name based on file name without extension
    table_name, ext = os.path.splitext(file_name)
    table_name = table_name.lower()
    # Ensure standard table name conventions
    table_name = re.sub(r"\W+", "_", table_name).strip("_")

    try:
        if ext.lower() in [".xlsx", ".xlsb"]:
            # Guard: check for HTML content
            with open(file_path, "rb") as f:
                header = f.read(500)
                if b"<!DOCTYPE" in header.upper() or b"<HTML" in header.upper():
                    raise ValueError(
                        f"File {file_name} appears to be an HTML error page, not a valid Excel file."
                    )

            last_sheet = None
            if ext.lower() == ".xlsx":
                engine_name = "openpyxl"
                # Load workbook in read-only mode to get sheet names quickly
                wb = load_workbook(file_path, read_only=True)
                last_sheet = wb.sheetnames[-1]
                wb.close()
            else:
                engine_name = "pyxlsb"
                # Access sheet names for binary excel files
                with open_xlsb(file_path) as wb:
                    last_sheet = wb.sheets[-1]  # wb.sheets returns list of sheet names

            # Pass the last sheet name to your reading function
            df = read_excel_with_dynamic_header(
                file_path, engine=engine_name, sheet_name=last_sheet
            )
        else:
            df = pd.read_csv(file_path)

        df = format_columns(df)

        # Add loaded_at timestamp for traceability
        df["loaded_at"] = pd.Timestamp.now()

        # Log DataFrame info before committing to Postgres
        logger.debug(
            f"Data Profile for {table_name}:\nColumns and Data Types:\n{df.dtypes}\n\nFirst 5 rows:\n{df.head(5)}\n"
        )

        # Write data to the 'bronze' schema
        df.to_sql(
            name=table_name,
            con=engine,
            schema="bronze",
            if_exists="replace",
            index=False,
        )
        logger.info(
            f"Task successfully completed: Loaded {len(df)} rows into bronze.{table_name}"
        )

    except Exception as e:
        logger.error(f"Error loading {file_name}: {e}")
        raise


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest SNIES file into PostgreSQL.")
    parser.add_argument(
        "--file", type=str, help="Absolute path to the file to process.", required=True
    )
    args = parser.parse_args()

    load_file(args.file)
